"""
Geração de planilha Excel a partir de um `CobrancaResultado`.

Usa openpyxl (já no requirements). Gera 1 aba "Registros" com 11 colunas
e cabeçalho bold + larguras razoáveis. Retorna bytes pra ser servido pelo
endpoint sem tocar disco.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from api.cobrancas.schema import CobrancaResultado


# Ordem das colunas + largura aproximada em chars.
COLUMN_SPECS: list[tuple[str, int]] = [
    ("CONDOMINIO", 36),
    ("UNIDADE", 10),
    ("PRIMEIRO_VENCTO", 14),
    ("MULTA", 10),
    ("EMISSAO", 12),
    ("NR_DO_RECIBO", 14),
    ("REGISTRO_EMISSAO", 16),
    ("SITUACAO", 14),
    ("CONTA", 8),
    ("HISTORICO", 36),
    ("VALOR_ORIGINAL", 14),
]


def gerar_xlsx(resultado: CobrancaResultado, *, file_name: str | None = None) -> bytes:
    """Devolve o `.xlsx` como bytes — pronto pra StreamingResponse."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    # Linha 1: header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="CB1D40")  # cor primária Lello
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx, (col_name, width) in enumerate(COLUMN_SPECS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Linhas 2..N: dados
    for row_idx, registro in enumerate(resultado.registros, start=2):
        valores = registro.model_dump()
        for col_idx, (col_name, _w) in enumerate(COLUMN_SPECS, start=1):
            v = valores.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=v)
        # Formato moeda nas colunas numéricas (MULTA=4, VALOR_ORIGINAL=11)
        ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=11).number_format = "#,##0.00"

    # Linha de totais — só se há registros.
    if resultado.registros:
        last_data_row = 1 + len(resultado.registros)
        total_row = last_data_row + 1
        total_label = ws.cell(row=total_row, column=10, value="TOTAL")
        total_label.font = Font(bold=True)
        total_label.alignment = Alignment(horizontal="right")
        total_value = ws.cell(
            row=total_row, column=11, value=resultado.metadata.total_valor
        )
        total_value.font = Font(bold=True)
        total_value.number_format = "#,##0.00"

    # Freeze pane no header pra rolar mantendo a 1ª linha visível.
    ws.freeze_panes = "A2"

    # Metadata — aba secundária, opcional. Útil pra rastreabilidade.
    if file_name or resultado.metadata.periodo or resultado.metadata.data_emissao_relatorio:
        meta = wb.create_sheet("Metadata")
        meta.column_dimensions["A"].width = 28
        meta.column_dimensions["B"].width = 50
        linhas = [
            ("Arquivo origem", file_name or ""),
            ("Período", resultado.metadata.periodo),
            ("Data emissão relatório", resultado.metadata.data_emissao_relatorio),
            ("Total registros", resultado.metadata.total_registros),
            ("Total valor", resultado.metadata.total_valor),
        ]
        for i, (k, v) in enumerate(linhas, start=1):
            meta.cell(row=i, column=1, value=k).font = Font(bold=True)
            meta.cell(row=i, column=2, value=v)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
