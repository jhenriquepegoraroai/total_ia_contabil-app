"""
ExcelFolderConnector — lê arquivos .xlsx/.xls de uma pasta e produz `RawChunk`s.

Estratégia:
  - Cada arquivo Excel é uma fonte de linhas.
  - Por default lê a primeira sheet; pode ser configurado.
  - Cada linha vira 1 RawChunk com `paragraph` = valor de `coluna_texto`.
  - `referencia` (opcional) vem de `coluna_referencia`; senão usa
    `referencia_default` setado na fonte (e o pipeline cai no padrão).
  - `data_valida` (opcional) vem de `coluna_data`.
  - `record_id` = "row<N>" baseado na linha (estável entre execuções).
"""

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

from loguru import logger

from .base import Connector, RawChunk


_EXCEL_RE = re.compile(r"\.(xlsx|xlsm|xls)$", re.IGNORECASE)


class ExcelFolderConnector(Connector):
    """
    Lê arquivos Excel de uma pasta (não-recursivo por default).

    Args:
        path: pasta com arquivos Excel.
        coluna_texto: nome da coluna a indexar (obrigatório).
        coluna_referencia: nome da coluna que vira `referencia` (opcional).
        coluna_data: nome da coluna de data (opcional).
        referencia_default: usado se a linha não trouxer referência.
        sheet_name: nome da aba (None = primeira aba).
        recursive: varrer subpastas também.
    """

    def __init__(
        self,
        path: str,
        *,
        coluna_texto: str,
        coluna_referencia: str | None = None,
        coluna_data: str | None = None,
        referencia_default: str | None = None,
        sheet_name: str | None = None,
        recursive: bool = False,
    ):
        self._path = Path(path)
        if not self._path.exists():
            raise FileNotFoundError(f"Pasta não existe: {self._path}")
        if not self._path.is_dir():
            raise NotADirectoryError(f"Path não é diretório: {self._path}")

        if not coluna_texto:
            raise ValueError("ExcelFolderConnector exige `coluna_texto`.")

        self._coluna_texto = coluna_texto
        self._coluna_ref = coluna_referencia
        self._coluna_data = coluna_data
        self._referencia_default = referencia_default
        self._sheet_name = sheet_name
        self._recursive = recursive

    def describe(self) -> str:
        return f"excel_folder:{self._path}"

    def read(self) -> Iterator[RawChunk]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl não instalado.") from exc

        pattern = "**/*" if self._recursive else "*"
        arquivos = sorted(p for p in self._path.glob(pattern) if _EXCEL_RE.search(p.name))
        if not arquivos:
            logger.warning(f"Nenhum arquivo Excel em {self._path}")
            return

        logger.info(f"[{self.describe()}] {len(arquivos)} planilhas")

        for arquivo in arquivos:
            try:
                yield from self._ler_arquivo(arquivo, load_workbook)
            except Exception as exc:
                logger.exception(f"Erro lendo {arquivo.name}: {exc}")
                continue

    def _ler_arquivo(self, arquivo: Path, load_workbook) -> Iterator[RawChunk]:
        wb = load_workbook(filename=str(arquivo), read_only=True, data_only=True)
        sheet = wb[self._sheet_name] if self._sheet_name else wb.active

        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            logger.warning(f"{arquivo.name}: planilha vazia.")
            wb.close()
            return

        if header is None:
            wb.close()
            return

        col_idx = {str(h).strip() if h is not None else "": i for i, h in enumerate(header)}

        if self._coluna_texto not in col_idx:
            logger.error(
                f"{arquivo.name}: coluna_texto '{self._coluna_texto}' não encontrada. "
                f"Colunas: {list(col_idx.keys())}"
            )
            wb.close()
            return

        idx_texto = col_idx[self._coluna_texto]
        idx_ref = col_idx.get(self._coluna_ref) if self._coluna_ref else None
        idx_data = col_idx.get(self._coluna_data) if self._coluna_data else None

        n_chunks = 0
        for n_row, row in enumerate(rows, start=2):  # row 2 = primeira linha de dados (1 = header)
            if not row:
                continue
            try:
                texto_val = row[idx_texto] if idx_texto < len(row) else None
                if texto_val is None:
                    continue
                texto = str(texto_val).strip()
                if not texto:
                    continue

                referencia: str | None = None
                if idx_ref is not None and idx_ref < len(row):
                    v = row[idx_ref]
                    if v is not None:
                        referencia = str(v).strip() or None
                if not referencia:
                    referencia = self._referencia_default

                data_valida: date | None = None
                if idx_data is not None and idx_data < len(row):
                    v = row[idx_data]
                    if isinstance(v, datetime):
                        data_valida = v.date()
                    elif isinstance(v, date):
                        data_valida = v

                yield RawChunk(
                    file_name=arquivo.name,
                    record_id=f"row{n_row}",
                    paragraph=texto,
                    data_valida=data_valida,
                    referencia=referencia,
                )
                n_chunks += 1

            except Exception as exc:
                logger.warning(f"{arquivo.name} linha {n_row}: {exc}")
                continue

        wb.close()
        logger.debug(f"[{arquivo.name}] {n_chunks} chunks extraídos")
